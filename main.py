import hashlib
import hmac
import io
import os
import re
import tempfile
import uuid
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from typing import List, Tuple, Optional

import pdfplumber
import psycopg
import requests
from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from psycopg.rows import dict_row
from pydantic import BaseModel

app = FastAPI(title="ExcelLimpio Backend v2", version="2.3.2")

FRONTEND_BASE_URL = os.getenv("FRONTEND_BASE_URL", "https://excel-limpio.netlify.app").rstrip("/")
MERCADO_PAGO_ACCESS_TOKEN = os.getenv("MERCADO_PAGO_ACCESS_TOKEN", "").strip()
MERCADO_PAGO_WEBHOOK_SECRET = os.getenv("MERCADO_PAGO_WEBHOOK_SECRET", "").strip()
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

PLANS = {
    "p3": {"name": "Hasta 3 páginas", "price": 9.90, "max_pages": 3, "max_size_mb": 5, "duration_days": 1, "one_time": True},
    "p10": {"name": "Hasta 10 páginas", "price": 24.90, "max_pages": 10, "max_size_mb": 10, "duration_days": 1, "one_time": True},
    "p25": {"name": "Hasta 25 páginas", "price": 39.90, "max_pages": 25, "max_size_mb": 20, "duration_days": 1, "one_time": True},
    "monthly": {"name": "Mensual 30 días", "price": 99.90, "max_pages": 300, "max_size_mb": 20, "duration_days": 30, "one_time": False},
}

thin_side = Side(style="thin", color="000000")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


class CheckoutRequest(BaseModel):
    plan: str


class ActivatePaymentRequest(BaseModel):
    payment_id: str
    purchase_id: str
    status_hint: Optional[str] = None


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat()


@contextmanager
def get_db():
    if not DATABASE_URL:
        raise RuntimeError("Falta configurar DATABASE_URL en Render.")
    conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    with get_db() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS purchases (
            id TEXT PRIMARY KEY,
            plan_code TEXT NOT NULL,
            price DOUBLE PRECISION NOT NULL,
            status TEXT NOT NULL,
            payment_id TEXT,
            created_at TEXT NOT NULL,
            activated_at TEXT
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS access_tokens (
            token TEXT PRIMARY KEY,
            purchase_id TEXT NOT NULL REFERENCES purchases(id),
            plan_code TEXT NOT NULL,
            plan_name TEXT NOT NULL,
            remaining_pages INTEGER,
            max_file_size_mb INTEGER NOT NULL,
            expires_at TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            used_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """)


@app.on_event("startup")
def on_startup():
    init_db()


def _cluster(vals: List[float], tol: float = 1.0) -> List[float]:
    if not vals:
        return []
    vals = sorted(vals)
    out: List[float] = []
    cur = [vals[0]]
    for v in vals[1:]:
        if abs(v - cur[-1]) <= tol:
            cur.append(v)
        else:
            out.append(sum(cur) / len(cur))
            cur = [v]
    out.append(sum(cur) / len(cur))
    return out


def _bbox_area(b: Tuple[float, float, float, float]) -> float:
    return max(0.0, (b[2] - b[0])) * max(0.0, (b[3] - b[1]))


def _bbox_contains(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float], margin: float = 2.0) -> bool:
    return (
        b[0] >= a[0] - margin
        and b[1] >= a[1] - margin
        and b[2] <= a[2] + margin
        and b[3] <= a[3] + margin
    )


def _iou(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float]) -> float:
    ix0 = max(a[0], b[0])
    iy0 = max(a[1], b[1])
    ix1 = min(a[2], b[2])
    iy1 = min(a[3], b[3])
    inter = max(0.0, ix1 - ix0) * max(0.0, iy1 - iy0)
    union = _bbox_area(a) + _bbox_area(b) - inter
    return inter / union if union > 0 else 0.0


def _table_grid_dims(table) -> Tuple[int, int]:
    xs: List[float] = []
    ys: List[float] = []
    for (x0, top, x1, bottom) in table.cells:
        xs.extend([x0, x1])
        ys.extend([top, bottom])
    xs = _cluster(xs, tol=1.0)
    ys = _cluster(ys, tol=1.0)
    return max(0, len(xs) - 1), max(0, len(ys) - 1)


def _points_to_excel_col_width(points: float) -> float:
    pixels = points * (96.0 / 72.0)
    width = max(2.0, (pixels - 5.0) / 7.0)
    return min(width, 80.0)


def _find_index(edges: List[float], value: float) -> int:
    best = 0
    best_d = 1e18
    for i, e in enumerate(edges):
        d = abs(e - value)
        if d < best_d:
            best_d = d
            best = i
    return best


def _extract_text_in_bbox(page, bbox: Tuple[float, float, float, float]) -> str:
    try:
        txt = page.crop(bbox).extract_text(x_tolerance=1, y_tolerance=1)
        if not txt:
            return ""
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines()]
        lines = [ln for ln in lines if ln]
        return "\n".join(lines)
    except Exception:
        return ""


def _extract_tables_candidates(page) -> List:
    settings_variants = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines", "intersection_tolerance": 5},
        {"vertical_strategy": "lines", "horizontal_strategy": "text", "intersection_tolerance": 5},
        {"vertical_strategy": "text", "horizontal_strategy": "lines", "intersection_tolerance": 5},
        {"vertical_strategy": "text", "horizontal_strategy": "text", "intersection_tolerance": 5},
    ]
    seen = set()
    out = []
    for settings in settings_variants:
        try:
            tables = page.find_tables(table_settings=settings)
        except Exception:
            continue
        for table in tables:
            bbox = tuple(round(x, 1) for x in table.bbox)
            if bbox in seen:
                continue
            if (bbox[2] - bbox[0]) < 50 or (bbox[3] - bbox[1]) < 30:
                continue
            seen.add(bbox)
            out.append(table)
    out.sort(key=lambda t: (t.bbox[1], t.bbox[0]))
    return out


def _extract_tables_filtered(page) -> List:
    candidates = _extract_tables_candidates(page)
    if not candidates:
        return []
    page_area = page.width * page.height
    enriched = []
    for table in candidates:
        area = _bbox_area(table.bbox)
        ncols, nrows = _table_grid_dims(table)
        enriched.append((table, area, ncols, nrows))
    filtered = []
    for table, area, ncols, nrows in enriched:
        grid = ncols * nrows
        if area < 0.03 * page_area and grid < 16:
            continue
        filtered.append((table, area))
    if not filtered:
        tmax = max(enriched, key=lambda x: x[1])[0]
        return [tmax]
    filtered.sort(key=lambda x: x[1], reverse=True)
    kept = []
    for table, _area in filtered:
        bbox = table.bbox
        skip = False
        for kept_table in kept:
            if _bbox_contains(kept_table.bbox, bbox, margin=1) and _iou(kept_table.bbox, bbox) > 0.85:
                skip = True
                break
        if not skip:
            kept.append(table)
    kept.sort(key=lambda t: (t.bbox[1], t.bbox[0]))
    return kept


def _add_original_page_sheet(
    wb: Workbook,
    page,
    page_index: int,
    temp_image_paths: List[str],
    dpi: int = 150
) -> None:
    ws = wb.create_sheet(title=f"ORIGINAL_p{page_index}")
    ws.sheet_view.showGridLines = False

    pil = page.to_image(resolution=dpi).original

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f:
        pil.save(f.name, format="PNG")
        img_path = f.name

    temp_image_paths.append(img_path)

    xlimg = XLImage(img_path)
    xlimg.anchor = "A1"
    ws.add_image(xlimg)

    ws.column_dimensions["A"].width = 2
    rows_needed = int(pil.height / 20) + 5
    for r in range(1, rows_needed + 1):
        ws.row_dimensions[r].height = 15


def _add_table_sheet(wb: Workbook, page, table, sheet_name: str, style_header: bool = True) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False

    xs: List[float] = []
    ys: List[float] = []
    for (x0, top, x1, bottom) in table.cells:
        xs.extend([x0, x1])
        ys.extend([top, bottom])

    xs = sorted(_cluster(xs, tol=1.0))
    ys = sorted(_cluster(ys, tol=1.0))

    if len(xs) < 2 or len(ys) < 2:
        return

    col_pts = [xs[i + 1] - xs[i] for i in range(len(xs) - 1)]
    row_pts = [ys[j + 1] - ys[j] for j in range(len(ys) - 1)]

    for i, wpt in enumerate(col_pts):
        ws.column_dimensions[get_column_letter(1 + i)].width = _points_to_excel_col_width(wpt)

    for j, hpt in enumerate(row_pts):
        ws.row_dimensions[1 + j].height = max(6.0, min(hpt, 409.0))

    wrap = Alignment(wrap_text=True, vertical="top")

    for r in range(1, len(ys)):
        for c in range(1, len(xs)):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = thin_border

    # Escribir texto sin combinar celdas para evitar error con MergedCell
    for bbox in table.cells:
        x0, top, x1, bottom = bbox
        c0 = _find_index(xs, x0)
        r0 = _find_index(ys, top)

        start_col = c0 + 1
        start_row = r0 + 1

        txt = _extract_text_in_bbox(page, bbox)
        if txt:
            try:
                cell = ws.cell(row=start_row, column=start_col)
                cell.value = txt
                cell.alignment = wrap
                cell.border = thin_border
            except Exception:
                pass

    if style_header:
        vals = [ws.cell(row=1, column=c).value for c in range(1, len(xs))]
        nonempty = sum(1 for v in vals if v and str(v).strip())
        if nonempty >= max(2, (len(xs) - 1) // 2):
            fill = PatternFill("solid", fgColor="0F172A")
            font = Font(bold=True, color="FFFFFF")
            center = Alignment(wrap_text=True, vertical="center", horizontal="center")
            for c in range(1, len(xs)):
                cell = ws.cell(row=1, column=c)
                cell.fill = fill
                cell.font = font
                cell.alignment = center


def _add_text_fallback_sheet(wb: Workbook, page, sheet_name: str) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False
    txt = page.extract_text() or ""
    lines = [ln.rstrip() for ln in txt.splitlines() if ln.strip()]
    ws.column_dimensions["A"].width = 120
    for i, ln in enumerate(lines[:2000], start=1):
        ws.cell(row=i, column=1, value=ln).alignment = Alignment(wrap_text=True, vertical="top")


def convert_pdf_to_xlsx_bytes(pdf_path: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    temp_image_paths: List[str] = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                _add_original_page_sheet(wb, page, i, temp_image_paths, dpi=150)

            for p_idx, page in enumerate(pdf.pages, start=1):
                tables = _extract_tables_filtered(page)
                if not tables:
                    _add_text_fallback_sheet(wb, page, sheet_name=f"EDITABLE_p{p_idx}_text")
                    continue

                for t_idx, table in enumerate(tables, start=1):
                    _add_table_sheet(
                        wb,
                        page,
                        table,
                        sheet_name=f"EDITABLE_p{p_idx}_t{t_idx}",
                        style_header=True
                    )

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()

    finally:
        for img_path in temp_image_paths:
            try:
                os.remove(img_path)
            except Exception:
                pass


def get_plan(plan_code: str):
    plan = PLANS.get(plan_code)
    if not plan:
        raise HTTPException(status_code=400, detail="Plan inválido.")
    return plan


def mp_headers():
    if not MERCADO_PAGO_ACCESS_TOKEN:
        raise HTTPException(status_code=500, detail="Falta configurar MERCADO_PAGO_ACCESS_TOKEN en Render.")
    return {
        "Authorization": f"Bearer {MERCADO_PAGO_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }


def create_purchase(plan_code: str) -> str:
    plan = get_plan(plan_code)
    purchase_id = str(uuid.uuid4())
    with get_db() as conn:
        conn.execute(
            "INSERT INTO purchases (id, plan_code, price, status, created_at) VALUES (%s, %s, %s, %s, %s)",
            (purchase_id, plan_code, plan["price"], "pending", iso(utcnow()))
        )
    return purchase_id


def create_checkout_preference(plan_code: str, purchase_id: str) -> str:
    plan = get_plan(plan_code)

    success_url = f"{FRONTEND_BASE_URL}/activate.html?purchase_id={purchase_id}"
    pending_url = f"{FRONTEND_BASE_URL}/activate.html?purchase_id={purchase_id}"
    failure_url = f"{FRONTEND_BASE_URL}/index.html?payment=failed&purchase_id={purchase_id}"

    payload = {
        "items": [
            {
                "title": f"ExcelLimpio - {plan['name']}",
                "quantity": 1,
                "currency_id": "PEN",
                "unit_price": plan["price"],
            }
        ],
        "back_urls": {
            "success": success_url,
            "pending": pending_url,
            "failure": failure_url,
        },
        "auto_return": "approved",
        "external_reference": purchase_id,
        "statement_descriptor": "EXCELLIMPIO",
        "binary_mode": True,
        "payment_methods": {
            "excluded_payment_types": [
                {"id": "ticket"},
                {"id": "atm"},
                {"id": "bank_transfer"}
            ],
            "installments": 1
        }
    }

    response = requests.post(
        "https://api.mercadopago.com/checkout/preferences",
        headers=mp_headers(),
        json=payload,
        timeout=30
    )

    if response.status_code >= 300:
        raise HTTPException(
            status_code=500,
            detail=f"Mercado Pago error creando checkout: {response.text}"
        )

    data = response.json()
    checkout_url = data.get("init_point") or data.get("sandbox_init_point")

    if not checkout_url:
        raise HTTPException(
            status_code=500,
            detail="Mercado Pago no devolvió init_point."
        )

    return checkout_url


def get_payment(payment_id: str) -> dict:
    response = requests.get(
        f"https://api.mercadopago.com/v1/payments/{payment_id}",
        headers={"Authorization": f"Bearer {MERCADO_PAGO_ACCESS_TOKEN}"},
        timeout=30
    )
    if response.status_code >= 300:
        raise HTTPException(status_code=400, detail=f"No se pudo validar el pago: {response.text}")
    return response.json()


def create_access_token_for_purchase(purchase_id: str, payment_id: str) -> dict:
    with get_db() as conn:
        purchase = conn.execute(
            "SELECT * FROM purchases WHERE id = %s",
            (purchase_id,)
        ).fetchone()

        if not purchase:
            raise HTTPException(status_code=404, detail="Compra no encontrada.")

        if purchase["status"] == "paid":
            existing = conn.execute(
                "SELECT * FROM access_tokens WHERE purchase_id = %s AND active = 1 ORDER BY created_at DESC LIMIT 1",
                (purchase_id,)
            ).fetchone()
            if existing:
                return existing

        plan = get_plan(purchase["plan_code"])
        token = str(uuid.uuid4())
        created_at = utcnow()
        expires_at = created_at + timedelta(days=plan["duration_days"])
        remaining_pages = plan["max_pages"]

        conn.execute("""
            INSERT INTO access_tokens
            (token, purchase_id, plan_code, plan_name, remaining_pages, max_file_size_mb, expires_at, active, used_count, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 1, 0, %s)
        """, (
            token,
            purchase_id,
            purchase["plan_code"],
            plan["name"],
            remaining_pages,
            plan["max_size_mb"],
            iso(expires_at),
            iso(created_at)
        ))

        conn.execute(
            "UPDATE purchases SET status = %s, payment_id = %s, activated_at = %s WHERE id = %s",
            ("paid", payment_id, iso(created_at), purchase_id)
        )

        row = conn.execute(
            "SELECT * FROM access_tokens WHERE token = %s",
            (token,)
        ).fetchone()

    return row


def validate_payment_for_purchase(payment: dict, purchase_id: str) -> None:
    status = payment.get("status")
    external_reference = str(payment.get("external_reference") or "")
    amount = float(payment.get("transaction_amount") or 0)

    with get_db() as conn:
        purchase = conn.execute(
            "SELECT * FROM purchases WHERE id = %s",
            (purchase_id,)
        ).fetchone()

    if not purchase:
        raise HTTPException(status_code=404, detail="Compra no encontrada.")

    expected_amount = float(purchase["price"])

    if status != "approved":
        raise HTTPException(status_code=400, detail="El pago todavía no está aprobado.")

    if external_reference != purchase_id:
        raise HTTPException(status_code=400, detail="El pago no coincide con la compra esperada.")

    if round(amount, 2) != round(expected_amount, 2):
        raise HTTPException(status_code=400, detail="El monto pagado no coincide con el plan.")


def activate_purchase_from_payment_id(payment_id: str):
    payment = get_payment(payment_id)
    purchase_id = str(payment.get("external_reference") or "").strip()

    if not purchase_id:
        return None

    validate_payment_for_purchase(payment, purchase_id)
    return create_access_token_for_purchase(purchase_id, payment_id)


def parse_signature_header(signature_header: str) -> dict:
    parts = [p.strip() for p in signature_header.split(",") if p.strip()]
    data = {}
    for part in parts:
        if "=" in part:
            k, v = part.split("=", 1)
            data[k.strip()] = v.strip()
    return data


def is_valid_mercadopago_signature(
    request: Request,
    x_signature: Optional[str],
    x_request_id: Optional[str],
    data_id: str
) -> bool:
    if not MERCADO_PAGO_WEBHOOK_SECRET:
        return False
    if not x_signature or not x_request_id or not data_id:
        return False

    parsed = parse_signature_header(x_signature)
    ts = parsed.get("ts")
    v1 = parsed.get("v1")

    if not ts or not v1:
        return False

    manifest = f"id:{data_id};request-id:{x_request_id};ts:{ts};"
    expected = hmac.new(
        MERCADO_PAGO_WEBHOOK_SECRET.encode("utf-8"),
        manifest.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()

    return hmac.compare_digest(expected, v1)


def parse_auth_token(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Falta Authorization Bearer token.")
    parts = authorization.split(" ", 1)
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=401, detail="Authorization inválido.")
    return parts[1].strip()


def get_active_access(token: str) -> dict:
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM access_tokens WHERE token = %s AND active = 1",
            (token,)
        ).fetchone()

    if not row:
        raise HTTPException(status_code=401, detail="Acceso inválido o inactivo.")

    expires_at = datetime.fromisoformat(row["expires_at"])

    if utcnow() > expires_at:
        with get_db() as conn:
            conn.execute("UPDATE access_tokens SET active = 0 WHERE token = %s", (token,))
        raise HTTPException(status_code=401, detail="El acceso ya venció.")

    if row["remaining_pages"] is not None and row["remaining_pages"] <= 0:
        with get_db() as conn:
            conn.execute("UPDATE access_tokens SET active = 0 WHERE token = %s", (token,))
        raise HTTPException(status_code=401, detail="El acceso ya no tiene páginas disponibles.")

    return row


def count_pdf_pages_and_validate(data: bytes, max_size_mb: int) -> int:
    max_bytes = max_size_mb * 1024 * 1024
    if len(data) > max_bytes:
        raise HTTPException(status_code=400, detail=f"Este acceso solo acepta archivos de hasta {max_size_mb} MB.")
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            page_count = len(pdf.pages)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el PDF. Verifique que el archivo no esté dañado.")
    return page_count


@app.get("/health")
def health():
    return {
        "status": "ok",
        "message": "ExcelLimpio backend activo",
        "db_mode": "postgres",
        "webhook_secret_configured": bool(MERCADO_PAGO_WEBHOOK_SECRET),
        "plans": {
            code: {
                "name": plan["name"],
                "price": plan["price"],
                "max_pages": plan["max_pages"],
                "max_size_mb": plan["max_size_mb"],
                "duration_days": plan["duration_days"],
            }
            for code, plan in PLANS.items()
        }
    }


@app.post("/api/create-checkout")
def api_create_checkout(payload: CheckoutRequest):
    plan = get_plan(payload.plan)
    purchase_id = create_purchase(payload.plan)
    checkout_url = create_checkout_preference(payload.plan, purchase_id)
    return {
        "purchase_id": purchase_id,
        "plan_code": payload.plan,
        "plan_name": plan["name"],
        "checkout_url": checkout_url,
    }


@app.post("/api/activate-payment")
def api_activate_payment(payload: ActivatePaymentRequest):
    payment = get_payment(payload.payment_id)
    validate_payment_for_purchase(payment, payload.purchase_id)
    token_row = create_access_token_for_purchase(payload.purchase_id, payload.payment_id)
    return {
        "access_token": token_row["token"],
        "plan_name": token_row["plan_name"],
        "expires_at": token_row["expires_at"],
        "remaining_pages": token_row["remaining_pages"],
        "max_file_size_mb": token_row["max_file_size_mb"],
    }


@app.post("/webhooks/mercadopago")
async def mercado_pago_webhook(
    request: Request,
    x_signature: Optional[str] = Header(default=None),
    x_request_id: Optional[str] = Header(default=None),
):
    try:
        body = await request.json()
    except Exception:
        body = {}

    payment_id = (
        request.query_params.get("data.id")
        or request.query_params.get("id")
        or ((body.get("data") or {}).get("id") if isinstance(body, dict) else None)
        or (body.get("id") if isinstance(body, dict) else None)
    )

    action = (
        request.query_params.get("type")
        or (body.get("type") if isinstance(body, dict) else None)
        or (body.get("action") if isinstance(body, dict) else None)
        or ""
    )

    if action and "payment" not in str(action):
        return {
            "received": True,
            "ignored": True,
            "reason": "not a payment event",
            "action": action
        }

    if not payment_id:
        return {
            "received": True,
            "ignored": True,
            "reason": "payment_id not found"
        }

    if not is_valid_mercadopago_signature(
        request=request,
        x_signature=x_signature,
        x_request_id=x_request_id,
        data_id=str(payment_id)
    ):
        raise HTTPException(status_code=401, detail="Firma inválida del webhook.")

    try:
        activate_purchase_from_payment_id(str(payment_id))
        return {
            "received": True,
            "processed": True,
            "payment_id": str(payment_id)
        }
    except HTTPException as e:
        return {
            "received": True,
            "processed": False,
            "payment_id": str(payment_id),
            "detail": str(e.detail)
        }


@app.get("/api/session")
def api_session(authorization: Optional[str] = Header(default=None)):
    token = parse_auth_token(authorization)
    row = get_active_access(token)
    return {
        "token": row["token"],
        "plan_code": row["plan_code"],
        "plan_name": row["plan_name"],
        "remaining_pages": row["remaining_pages"],
        "max_file_size_mb": row["max_file_size_mb"],
        "expires_at": row["expires_at"],
        "used_count": row["used_count"],
    }


@app.post("/convert")
async def convert(file: UploadFile = File(...), authorization: Optional[str] = Header(default=None)):
    token = parse_auth_token(authorization)
    access = get_active_access(token)

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se acepta PDF.")

    data = await file.read()
    if not data or len(data) < 1000:
        raise HTTPException(status_code=400, detail="PDF vacío o inválido.")

    page_count = count_pdf_pages_and_validate(data, access["max_file_size_mb"])
    remaining_pages = access["remaining_pages"]

    if remaining_pages is not None and page_count > remaining_pages:
        raise HTTPException(
            status_code=400,
            detail=f"Este acceso solo tiene {remaining_pages} páginas disponibles."
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
        f.write(data)
        pdf_path = f.name

    try:
        xlsx_bytes = convert_pdf_to_xlsx_bytes(pdf_path)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error convirtiendo PDF: {e}")
    finally:
        try:
            os.remove(pdf_path)
        except Exception:
            pass

    with get_db() as conn:
        new_remaining = remaining_pages - page_count if remaining_pages is not None else None
        new_used_count = int(access["used_count"]) + 1
        new_active = 1

        if access["plan_code"] in ("p3", "p10", "p25"):
            new_remaining = 0
            new_active = 0
        elif new_remaining is not None and new_remaining <= 0:
            new_remaining = 0
            new_active = 0

        conn.execute(
            "UPDATE access_tokens SET remaining_pages = %s, used_count = %s, active = %s WHERE token = %s",
            (new_remaining, new_used_count, new_active, token)
        )

    out_name = re.sub(r"\.pdf$", "", os.path.basename(file.filename), flags=re.I) + "_ExcelLimpio.xlsx"
    headers = {"Content-Disposition": f'attachment; filename=\"{out_name}\"'}

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
